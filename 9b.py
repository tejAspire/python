from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
sheet = wb.active
sheet.title = "Language"
wb.create_sheet(title="Capital")

lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ['KA', 'TS', 'TN']

# Set header row and apply bold font
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Language"
sheet.cell(row=1, column=3).value = "Code"
ft = Font(bold=True)

for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft

# Fill in data for the "Language" sheet
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = lang[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]

wb.save("demo.xlsx")

# Switch to the "Capital" sheet
sheet = wb["Capital"]

# Set header row and apply bold font
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Capital"
sheet.cell(row=1, column=3).value = "Code"
ft = Font(bold=True)

for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft

# Fill in data for the "Capital" sheet
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = capital[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]

wb.save("demo.xlsx")
